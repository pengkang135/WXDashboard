import html
import os
import re
import shutil
from datetime import datetime, timedelta

from .config import get_wx_file_path
from .database import get_db

DOWNLOAD_ROOT = r"F:\WXDashboard\.Download"

EXTERNAL_CATEGORIES = ["供应商咨询", "地基处理", "建筑MEP", "保险", "物流"]

SUB_CATEGORY_EN = {
    "钢管桩": "SteelPipePile",
    "管桩": "PipePile",
    "地基处理": "GroundImprovement",
    "桩基/地基": "PileFoundation",
    "建筑总承包": "BuildingContractor",
    "机电总包": "MEPContractor",
    "管道": "Pipe",
    "疏浚": "Dredging",
    "保险经纪": "InsuranceBroker",
    "保险服务": "InsuranceService",
    "物流": "Logistics",
    "石料/土工": "StoneGeotextile",
    "供油/润滑": "OilLubricant",
    "钢材": "Steel",
    "钢轨": "Rail",
    "钢结构": "SteelStructure",
    "橡胶护舷": "RubberFender",
    "护舷": "RubberFender",
    "泵/液压": "PumpHydraulic",
    "光伏设备": "SolarEquipment",
    "漂浮安装": "FloatingInstallation",
    "电气": "Electrical",
    "机电安装": "MEPInstallation",
    "其他": "Others",
}

PRICE_KEYWORDS = [
    "price", "rate", "amount", "unit price", "total", "cost", "quotation",
    "单价", "总价", "金额", "报价", "合计", "价格", "费用", "含税",
    "subtotal", "discount", "net price", "gross price", "FOB", "CIF", "EXW",
]

BROCHURE_KEYWORDS = [
    "brochure", "company profile", "introduction", "catalog",
    "介绍册", "公司简介", "产品目录", "公司介绍",
]


def _get_internal_senders(conn):
    """Dynamically identify our people: anyone who sent messages in internal groups."""
    rows = conn.execute("""
        SELECT DISTINCT m.sender
        FROM messages m
        JOIN groups g ON m.group_id = g.id
        WHERE g.category IN ('内部沟通', '施工局合作', '设计院合作')
    """).fetchall()
    return {r["sender"] for r in rows if r["sender"]}


def collect_file_messages(conn, days=7):
    """Return list of dicts: {id, group_id, sender, content, msg_time, msg_date,
    group_name, category, sub_category, filename, ext}"""
    since = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    internal_senders = _get_internal_senders(conn)

    rows = conn.execute("""
        SELECT m.id, m.group_id, m.sender, m.content, m.msg_time, m.msg_date,
               g.name AS group_name, g.category, g.sub_category
        FROM messages m
        JOIN groups g ON m.group_id = g.id
        WHERE m.msg_date >= ?
          AND m.content LIKE '%[文件]%'
          AND g.category IN ({})
        ORDER BY m.msg_time DESC
    """.format(",".join("?" * len(EXTERNAL_CATEGORIES))),
        [since] + list(EXTERNAL_CATEGORIES)
    ).fetchall()

    results = []
    for row in rows:
        # Extract filename from content: [文件] filename.ext
        m = re.search(r'\[文件\]\s*(.+?)(?:\n|$)', row["content"])
        if not m:
            continue
        filename = html.unescape(m.group(1).strip())
        ext_match = re.search(r'\.(pdf|xlsx)$', filename, re.IGNORECASE)
        if not ext_match:
            continue

        # Skip files from our own people
        if row["sender"] in internal_senders:
            continue

        results.append({
            "id": row["id"],
            "group_id": row["group_id"],
            "sender": row["sender"],
            "content": row["content"],
            "msg_time": row["msg_time"],
            "msg_date": row["msg_date"],
            "group_name": row["group_name"],
            "category": row["category"],
            "sub_category": row["sub_category"],
            "filename": filename,
            "ext": ext_match.group(1).lower(),
        })
    return results


def extract_company_from_group(group_name):
    """Extract supplier/company name from group name.
    'Laldia-CHEC&裕大管桩' -> '裕大管桩'
    '孟加拉Laldia项目-港湾-中岩大地' -> '中岩大地'
    'Laldia-MEP-CHEC&Inspur浪潮' -> 'Inspur_浪潮'
    """
    # Pattern 1: CHEC&XXX or CHEC&XXX中文
    m = re.search(r'CHEC&(.+?)$', group_name)
    if m:
        return m.group(1).strip()

    # Pattern 2: 项目-港湾-XXX or 项目-单位-XXX
    parts = group_name.split("-")
    if len(parts) >= 3:
        return parts[-1].strip()

    return group_name


def build_dir_name(sub_category, company):
    """Build bilingual directory name.
    sub_category='钢管桩', company='裕大管桩' -> '钢管桩_SteelPipePile/裕大管桩'
    """
    en = SUB_CATEGORY_EN.get(sub_category, sub_category)
    cat_dir = f"{sub_category}_{en}" if sub_category else f"未分类_Unclassified"
    company_dir = company.replace("/", "_").replace("\\", "_").strip()
    return os.path.join(cat_dir, company_dir)


def _xlsx_has_price(file_path):
    """Check if xlsx file contains price information."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(max_row=50, values_only=True):
                row_text = " ".join(str(c) for c in row if c is not None)
                row_lower = row_text.lower()
                if any(kw in row_lower for kw in PRICE_KEYWORDS):
                    wb.close()
                    return True
        wb.close()
    except Exception:
        pass
    return False


def _pdf_has_price(file_path):
    """Check if PDF file contains price information."""
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            pages_to_check = pdf.pages[:3]
            text = "\n".join(
                page.extract_text() or "" for page in pages_to_check
            )
        text_lower = text.lower()
        if any(kw in text_lower for kw in PRICE_KEYWORDS):
            return True
        if re.search(r'[\$\€\¥\£]', text):
            return True
    except Exception:
        pass
    return False


def is_price_document(file_path, ext):
    """Check if file is a quotation document."""
    if ext == "xlsx":
        return _xlsx_has_price(file_path)
    elif ext == "pdf":
        return _pdf_has_price(file_path)
    return False


def is_brochure(file_path, ext, filename):
    """Check if file appears to be a company brochure."""
    name_lower = filename.lower()
    if any(kw in name_lower for kw in BROCHURE_KEYWORDS):
        return True
    return False
