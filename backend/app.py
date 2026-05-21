import threading
import time
import subprocess
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Flask, render_template, request, jsonify, send_file
from .config import TEMPLATES_DIR, STATIC_DIR, FLASK_HOST, FLASK_PORT, get_wx_file_url, get_wx_file_path
from .database import (
    init_db, get_db, get_all_groups, get_group, get_categories,
    get_messages, get_latest_messages, search_messages,
    get_sync_status, get_message_count, get_contacts, get_projects,
    set_subcategory, update_group_category, update_group_settings,
    get_summaries, get_extractions
)
from .sync_engine import sync_incremental, sync_all_groups_full, sync_full, get_sync_stats, discover_new_groups

app = Flask(__name__, template_folder=TEMPLATES_DIR, static_folder=STATIC_DIR)

_auto_sync_thread = None
_auto_sync_interval = 60
_auto_sync_running = False
_auto_sync_last_result = None
_last_heartbeat = 0

_manual_sync_running = False
_manual_sync_result = None


def _auto_sync_loop():
    global _auto_sync_running, _auto_sync_last_result, _last_heartbeat
    while _auto_sync_running:
        time.sleep(_auto_sync_interval)
        if not _auto_sync_running:
            break
        # 心跳超时(90秒)则自动停止
        if _last_heartbeat and time.time() - _last_heartbeat > 600:
            _auto_sync_running = False
            break
        try:
            _auto_sync_last_result = sync_incremental()
        except Exception as e:
            _auto_sync_last_result = {"error": str(e)}


@app.route("/")
def dashboard():
    return render_template("dashboard.html")


@app.route("/api/groups")
def api_groups():
    category = request.args.get("category")
    project = request.args.get("project", "Laldia")
    group_creator = request.args.get("group_creator")
    with_details = request.args.get("with_details", "0") == "1"
    groups = get_all_groups(category=category, project=project, group_creator=group_creator, with_details=with_details)
    return jsonify(groups)


@app.route("/api/groups/<int:group_id>")
def api_group_detail(group_id):
    group = get_group(group_id)
    if not group:
        return jsonify({"error": "群组不存在"}), 404
    group["message_count"] = get_message_count(group_id)
    return jsonify(group)


@app.route("/api/categories")
def api_categories():
    project = request.args.get("project", "Laldia")
    return jsonify(get_categories(project=project))


@app.route("/api/groups/<int:group_id>/messages")
def api_group_messages(group_id):
    limit = request.args.get("limit", 50, type=int)
    offset = request.args.get("offset", 0, type=int)
    messages, total = get_messages(group_id, limit=limit, offset=offset)
    return jsonify({"messages": messages, "total": total, "limit": limit, "offset": offset})


@app.route("/api/groups/<int:group_id>/messages/latest")
def api_group_latest_messages(group_id):
    n = request.args.get("n", 3, type=int)
    messages = get_latest_messages(group_id, n=n)
    return jsonify(messages)


@app.route("/api/search")
def api_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"error": "缺少搜索关键词"}), 400
    group_id = request.args.get("group_id", type=int)
    results = search_messages(q, group_id=group_id)
    return jsonify({"query": q, "results": results, "count": len(results)})


@app.route("/api/sync/refresh", methods=["POST"])
def api_sync_refresh():
    global _manual_sync_running, _manual_sync_result
    if _manual_sync_running:
        return jsonify({"status": "running", "message": "同步已在执行中"})
    _manual_sync_running = True
    _manual_sync_result = None
    def _run():
        global _manual_sync_running, _manual_sync_result
        try:
            _manual_sync_result = sync_incremental()
        except Exception as e:
            _manual_sync_result = {"error": str(e)}
        finally:
            _manual_sync_running = False
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"status": "started"})


@app.route("/api/sync/refresh/status")
def api_sync_refresh_status():
    return jsonify({
        "running": _manual_sync_running,
        "result": _manual_sync_result
    })


@app.route("/api/sync/pull-all", methods=["POST"])
def api_sync_pull_all():
    group_name = request.args.get("group")
    try:
        if group_name:
            result = sync_full(group_name)
            return jsonify(result)
        else:
            result = sync_all_groups_full()
            return jsonify(result)
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/sync/status")
def api_sync_status():
    stats = get_sync_stats()
    return jsonify(stats)


@app.route("/api/projects")
def api_projects():
    return jsonify(get_projects())


@app.route("/api/groups/<int:group_id>/contacts")
def api_group_contacts(group_id):
    contacts = get_contacts(group_id=group_id)
    return jsonify(contacts)


@app.route("/api/contacts/bcc")
def api_contacts_bcc():
    """Get unique contacts with emails for a category, for BCC mass mail."""
    category = request.args.get("category", "")
    project = request.args.get("project", "Laldia")
    if not category:
        return jsonify({"error": "缺少分类参数"}), 400

    from .config import MY_WECHAT_NAME, MY_EMAIL_KEYWORD
    conn = get_db()
    rows = conn.execute("""
        SELECT c.sender_name, c.email FROM contacts c
        JOIN groups g ON c.group_id = g.id
        WHERE g.deleted=0 AND g.category=? AND g.project=?
          AND c.email IS NOT NULL AND c.email != ''
          AND c.sender_name IS NOT ?
          AND (c.email NOT LIKE ?)
        ORDER BY c.sender_name
    """, (category, project, MY_WECHAT_NAME, "%" + MY_EMAIL_KEYWORD + "%")).fetchall()
    conn.close()

    seen = set()
    result = []
    for r in rows:
        key = r["email"].lower()
        if key in seen:
            continue
        seen.add(key)
        result.append({"name": r["sender_name"] or "", "email": r["email"]})
    return jsonify(result)


@app.route("/api/groups/<int:group_id>/summaries")
def api_group_summaries(group_id):
    summaries = get_summaries(group_id)
    return jsonify(summaries)


@app.route("/api/groups/<int:group_id>/extractions")
def api_group_extractions(group_id):
    extractions = get_extractions(group_id)
    return jsonify(extractions)


@app.route("/api/files/check")
def api_check_file():
    msg_date = request.args.get("msg_date", "")
    filename = request.args.get("filename", "")
    url = get_wx_file_url(msg_date, filename)
    if url:
        return jsonify({"exists": True, "url": url})
    return jsonify({"exists": False})


@app.route("/api/files/open", methods=["POST"])
def api_open_file():
    msg_date = request.args.get("msg_date", "")
    filename = request.args.get("filename", "")
    file_path = get_wx_file_path(msg_date, filename)
    if file_path:
        import os as _os
        subprocess.Popen(["explorer", "/select,", _os.path.normpath(file_path)])
        return jsonify({"ok": True})
    return jsonify({"ok": False, "error": "文件不存在"})


@app.route("/api/images/open", methods=["POST"])
def api_open_image():
    msg_date = request.args.get("msg_date", "")
    timestamp = request.args.get("timestamp")
    if timestamp:
        try:
            timestamp = int(timestamp)
        except (TypeError, ValueError):
            timestamp = None

    from .config import find_image_dat, decrypt_dat_file
    dat_files = find_image_dat(msg_date, timestamp=timestamp)

    if not dat_files:
        return jsonify({"ok": False, "error": "未找到对应图片文件"})

    for dat_path in dat_files[:10]:
        img_data, ext = decrypt_dat_file(dat_path)
        if img_data and ext:
            import os as _os
            import tempfile
            tmp = tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False)
            tmp.write(img_data)
            tmp.close()
            _os.startfile(tmp.name)
            return jsonify({"ok": True, "file": _os.path.basename(dat_path)})

    # 解密失败,打开文件所在目录
    if dat_files:
        import os as _os
        subprocess.Popen(["explorer", "/select,", _os.path.normpath(dat_files[0])])
        return jsonify({"ok": True, "opened": "explorer"})

    return jsonify({"ok": False, "error": "无法解密图片"})


@app.route("/api/images/view")
def api_view_image():
    """Serve a decrypted image for inline display. Uses timestamp-based matching."""
    msg_date = request.args.get("msg_date", "")
    timestamp = request.args.get("timestamp")
    local_id = request.args.get("local_id", type=int)
    group_id = request.args.get("group_id", type=int)
    if timestamp:
        try:
            timestamp = int(timestamp)
        except (TypeError, ValueError):
            timestamp = None

    from .config import find_image_dat, decrypt_dat_file
    dat_files = find_image_dat(msg_date, timestamp=timestamp)
    if local_id is not None:
        ordered = _chrono_match(msg_date, local_id, group_id)
        if ordered:
            dat_files = ordered + dat_files

    for dat_path in dat_files[:10]:
        img_data, ext = decrypt_dat_file(dat_path)
        if img_data and ext:
            mime = "image/" + ("jpeg" if ext == "jpg" else ext)
            return send_file(io.BytesIO(img_data), mimetype=mime)

    return jsonify({"ok": False, "error": "无法解密图片"}), 404


def _chrono_match(msg_date, local_id, group_id=None):
    """Match .dat file by timestamp proximity, scoped to a specific group."""
    import json
    import os as _os
    from .config import _WX_ATTACH_ROOT
    if not _WX_ATTACH_ROOT or not msg_date:
        return []
    yyyy_mm = msg_date[:7]

    target_ts = None
    try:
        conn = get_db()
        if group_id:
            row = conn.execute("""
                SELECT raw_json FROM messages
                WHERE group_id=? AND local_id=? AND msg_date LIKE ?
            """, (group_id, local_id, yyyy_mm + '%')).fetchone()
        else:
            row = conn.execute("""
                SELECT raw_json FROM messages
                WHERE local_id=? AND msg_date LIKE ?
            """, (local_id, yyyy_mm + '%')).fetchone()
        conn.close()
        if row:
            raw = json.loads(row['raw_json'] or '{}')
            target_ts = raw.get('timestamp')
    except Exception:
        pass

    if not target_ts:
        return []

    candidates = []
    try:
        for hash_dir in _os.listdir(_WX_ATTACH_ROOT):
            img_dir = _os.path.join(_WX_ATTACH_ROOT, hash_dir, yyyy_mm, "Img")
            if not _os.path.isdir(img_dir):
                continue
            for f in _os.listdir(img_dir):
                if not f.endswith('.dat'):
                    continue
                fpath = _os.path.join(img_dir, f)
                if '_t.dat' in f or '_h.dat' in f:
                    continue
                try:
                    file_ts = int(f.replace('.dat', ''))
                except ValueError:
                    file_ts = int(_os.path.getmtime(fpath))
                candidates.append((abs(file_ts - target_ts), fpath))
    except OSError:
        pass

    candidates.sort(key=lambda x: x[0])
    return [c[1] for c in candidates[:3]]


@app.route("/api/groups/<int:group_id>/subcategory", methods=["POST"])
def api_set_subcategory(group_id):
    data = request.get_json(force=True)
    sub = data.get("sub_category", "")
    set_subcategory(group_id, sub)
    return jsonify({"ok": True})


@app.route("/api/groups/<int:group_id>/settings", methods=["GET", "POST"])
def api_group_settings(group_id):
    if request.method == "GET":
        group = get_group(group_id)
        if not group:
            return jsonify({"error": "群组不存在"}), 404
        return jsonify({
            "project": group.get("project", ""),
            "category": group.get("category", ""),
            "sub_category": group.get("sub_category", ""),
            "manual_category": bool(group.get("manual_category", 0)),
            "projects": get_projects(),
            "categories": get_categories()
        })
    data = request.get_json(force=True)
    if data.get("unlock"):
        from .database import unlock_group_category
        unlock_group_category(group_id)
        return jsonify({"ok": True})
    project = data.get("project", "")
    category = data.get("category", "")
    sub_category = data.get("sub_category", "")
    update_group_settings(group_id, project=project, category=category, sub_category=sub_category)
    return jsonify({"ok": True})


@app.route("/api/groups/<int:group_id>/category", methods=["POST"])
def api_set_category(group_id):
    data = request.get_json(force=True)
    cat = data.get("category", "")
    if not cat:
        return jsonify({"error": "分类不能为空"}), 400
    update_group_category(group_id, cat)
    return jsonify({"ok": True, "category": cat})


@app.route("/api/sync/discover", methods=["POST"])
def api_sync_discover():
    try:
        new_groups = discover_new_groups()
        return jsonify({"new_groups": new_groups, "count": len(new_groups)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/heartbeat", methods=["POST"])
def api_heartbeat():
    global _last_heartbeat
    _last_heartbeat = time.time()
    return jsonify({"ok": True})


@app.route("/api/sync/auto/status")
def api_auto_sync_status():
    return jsonify({
        "running": _auto_sync_running,
        "interval": _auto_sync_interval,
        "last_result": _auto_sync_last_result
    })


@app.route("/api/sync/auto/start", methods=["POST"])
def api_auto_sync_start():
    global _auto_sync_thread, _auto_sync_running, _auto_sync_interval
    data = request.get_json(silent=True) or {}
    interval = int(data.get("interval", 60))
    if interval < 120:
        interval = 120
    _auto_sync_interval = interval

    if _auto_sync_running:
        return jsonify({"running": True, "interval": _auto_sync_interval})

    _auto_sync_running = True
    _auto_sync_thread = threading.Thread(target=_auto_sync_loop, daemon=True)
    _auto_sync_thread.start()
    return jsonify({"running": True, "interval": _auto_sync_interval})


@app.route("/api/sync/auto/stop", methods=["POST"])
def api_auto_sync_stop():
    global _auto_sync_running
    _auto_sync_running = False
    return jsonify({"running": False})


@app.route("/api/export/excel")
def api_export_excel():
    project = request.args.get("project", "Laldia")
    category = request.args.get("category")
    group_creator = request.args.get("group_creator")
    groups = get_all_groups(category=category, project=project, group_creator=group_creator)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "微信群台账"

    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1F25", end_color="1A1F25", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D5DCE4"),
        right=Side(style="thin", color="D5DCE4"),
        top=Side(style="thin", color="D5DCE4"),
        bottom=Side(style="thin", color="D5DCE4"),
    )
    cell_font = Font(name="微软雅黑", size=9, color="3A4454")
    cell_align = Alignment(vertical="top", wrap_text=True)
    name_font = Font(name="微软雅黑", size=9, bold=True, color="1A1F25")
    time_font = Font(name="Consolas", size=9, color="6B7B8D")

    headers = ["#", "群名", "分类", "子分类", "最后活跃", "消息数", "群主", "最近消息", "联系信息"]
    col_widths = [4, 28, 12, 12, 18, 8, 14, 55, 28]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col_widths[ci - 1]

    for ri, g in enumerate(groups, 1):
        gid = g["id"]
        idx = ri
        name = g.get("name", "")
        cat = g.get("category", "")
        sub = g.get("sub_category", "")
        last_active = g.get("last_active_date", "")
        msg_count = get_message_count(gid)
        creator = g.get("group_creator", "")

        msgs = get_latest_messages(gid, 3)
        msgs_text = ""
        for m in msgs:
            time_part = m["msg_date"] if m["msg_date"] else ""
            sender = m["sender"] if m["sender"] else ""
            content = m["content"] if m["content"] else ""
            content = content.replace("\n", " ").replace("\r", "")
            if len(content) > 120:
                content = content[:120] + "..."
            msgs_text += f"{time_part} {sender}: {content}\n"
        msgs_text = msgs_text.rstrip("\n")

        contacts = get_contacts(group_id=gid)
        contacts_text = "\n".join(
            f"{c['sender_name']} / {c['email']}" if c.get("email") else c.get("sender_name", "")
            for c in contacts[:8]
        )

        row_data = [idx, name, cat, sub, last_active, msg_count, creator, msgs_text, contacts_text]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri + 1, column=ci, value=val if val is not None else "")
            cell.font = name_font if ci == 2 else cell_font
            if ci == 5:
                cell.font = time_font
            cell.alignment = cell_align
            cell.border = thin_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    import datetime
    today = datetime.date.today().strftime("%Y%m%d")
    cat_part = category if category else "全部"
    filename = f"WXGLedger_{project}_{cat_part}_{today}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    init_db()
    print(f"Laldia 微信群台账启动: http://{FLASK_HOST}:{FLASK_PORT}")
    app.run(host=FLASK_HOST, port=FLASK_PORT, debug=True)
