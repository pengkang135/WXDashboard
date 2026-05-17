import json
import os
from openpyxl import load_workbook
from .config import LEGACY_EXCEL, LEGACY_WX_JSON_DIR
from .database import init_db, upsert_group, upsert_message, update_group_stats, get_all_group_names


def import_groups_from_excel():
    wb = load_workbook(LEGACY_EXCEL, read_only=True)
    ws = wb["微信群台账"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, name, category, partner, purpose, _, last_active, notes, creator, contact = row
        if not name:
            continue
        upsert_group(
            name=str(name).strip(),
            category=str(category).strip() if category else "未分类",
            partner_full_name=str(partner).strip() if partner else None,
            main_purpose=str(purpose).strip() if purpose else None,
            notes=str(notes).strip() if notes else None,
            group_creator=str(creator).strip() if creator else None,
        )
        count += 1
    wb.close()
    print(f"[import] 从 Excel 导入 {count} 个群组")


def match_group_name(filename_no_ext):
    group_names = get_all_group_names()
    if filename_no_ext in group_names:
        return filename_no_ext
    replaced = filename_no_ext.replace("_", "&")
    if replaced in group_names:
        return replaced
    replaced2 = filename_no_ext.replace("&", "_")
    if replaced2 in group_names:
        return replaced2
    for gname in group_names:
        if gname.replace(" ", "") == filename_no_ext.replace(" ", ""):
            return gname
    for gname in group_names:
        if filename_no_ext in gname or gname in filename_no_ext:
            return gname
    return None


def import_messages_from_wx_json():
    if not os.path.isdir(LEGACY_WX_JSON_DIR):
        print(f"[import] wx_json 目录不存在: {LEGACY_WX_JSON_DIR}")
        return

    group_names = get_all_group_names()
    total_imported = 0

    for fname in os.listdir(LEGACY_WX_JSON_DIR):
        if not fname.endswith(".json"):
            continue

        fpath = os.path.join(LEGACY_WX_JSON_DIR, fname)
        candidate = fname[:-5]

        matched = match_group_name(candidate)
        if not matched:
            print(f"[import] 跳过 {fname}：无法匹配到群组（候选名: {candidate}）")
            continue

        from .database import get_db
        conn = get_db()
        group = conn.execute("SELECT id FROM groups WHERE name=?", (matched,)).fetchone()
        conn.close()
        if not group:
            print(f"[import] 跳过 {fname}：群组 '{matched}' 不在数据库中")
            continue

        group_id = group["id"]

        with open(fpath, "r", encoding="utf-8") as f:
            raw_text = f.read().strip()
        if not raw_text.startswith("["):
            print(f"[import] 跳过 {fname}：不是 JSON 数组")
            continue
        messages = json.loads(raw_text)

        count = 0
        for msg in messages:
            local_id = msg.get("local_id")
            sender = msg.get("sender", "未知")
            content = msg.get("content", "")
            msg_time = msg.get("time", "")
            msg_date = msg_time[:10] if msg_time else ""
            msg_type = msg.get("type", "text")
            raw_json = json.dumps(msg, ensure_ascii=False)

            upsert_message(group_id, local_id, sender, content, msg_time, msg_date,
                          msg_type, raw_json)
            count += 1

        last_date = messages[-1]["time"][:10] if messages and messages[-1].get("time") else None
        update_group_stats(group_id, last_active_date=last_date, total_messages=count)
        print(f"[import] {matched}: {count} 条消息, 最后活跃 {last_date}")
        total_imported += count

    print(f"[import] 总计导入 {total_imported} 条消息")


def verify_import():
    from .database import get_db
    conn = get_db()
    group_count = conn.execute("SELECT COUNT(*) as cnt FROM groups WHERE deleted=0").fetchone()["cnt"]
    msg_count = conn.execute("SELECT COUNT(*) as cnt FROM messages").fetchone()["cnt"]
    conn.close()
    print(f"[verify] 群组: {group_count}, 消息: {msg_count}")
    if group_count == 26:
        print("[verify] 群组数量匹配 (26)")
    else:
        print(f"[verify] WARNING: 预期 26 个群组, 实际 {group_count}")


if __name__ == "__main__":
    print("[import] 初始化数据库...")
    init_db()
    print("[import] 导入群组...")
    import_groups_from_excel()
    print("[import] 导入消息...")
    import_messages_from_wx_json()
    print("[import] 验证...")
    verify_import()
    print("[import] 完成")
